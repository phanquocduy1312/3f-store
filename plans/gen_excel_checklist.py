"""
3F Store – Excel Checklist Generator
Tạo file Excel chỉnh chu, đẹp, đúng nghiệp vụ cho dự án 3F Store & 3F Club.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_TEXT
import os
from datetime import datetime

OUTPUT_PATH = r"c:\Users\Admin\Downloads\ccc\plans\3F-Store-Feature-Checklist-2026.xlsx"
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# ─── Color Palette ──────────────────────────────────────────────────────────
C_BRAND_DARK   = "0B1F3A"   # Tiêu đề chính
C_BRAND_BLUE   = "0057E7"   # Header cột
C_BRAND_LIGHT  = "EEF4FF"   # Nền header nhẹ
C_WHITE        = "FFFFFF"
C_ROW_ALT      = "F8FBFF"   # Hàng chẵn
C_DONE_BG      = "ECFDF3"   # Xong – xanh lá nhẹ
C_DONE_TEXT    = "027A48"
C_PENDING_BG   = "FFF4E8"   # Đang làm – cam nhẹ
C_PENDING_TEXT = "B54708"
C_TODO_BG      = "FFF1F3"   # Chưa – đỏ nhẹ
C_TODO_TEXT    = "C01048"
C_SKIP_BG      = "F8FAFC"   # Bỏ qua – xám nhẹ
C_SKIP_TEXT    = "64748B"
C_HEADER_ROW   = "0057E7"   # Header xanh đậm
C_GROUP_BG     = "0B1F3A"   # Group header xanh navy
C_PRIO_P0      = "EF4444"   # P0 – Đỏ
C_PRIO_P1      = "F97316"   # P1 – Cam
C_PRIO_P2      = "6366F1"   # P2 – Tím

# ─── Border styles ──────────────────────────────────────────────────────────
thin  = Side(style="thin",   color="DCE8FF")
thick = Side(style="medium", color="0057E7")
no    = Side(style=None)

def border(top=thin, bottom=thin, left=thin, right=thin):
    return Border(top=top, bottom=bottom, left=left, right=right)

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def font(bold=False, color=C_BRAND_DARK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─── Feature Data ────────────────────────────────────────────────────────────
# Columns: group, id, feature, priority, status, owner, deadline, note
GROUPS = [
    {
        "title": "🛒  PHÂN HỆ MUA SẮM & THANH TOÁN",
        "color": "0057E7",
        "rows": [
            ("A4.1", "Tìm kiếm sản phẩm real-time + dropdown gợi ý",  "P0", "✅ Hoàn thành", "Frontend",    "",           "Debounce 300ms, map full product catalog"),
            ("A4.2", "Trang chi tiết sản phẩm (PDP) đầy đủ",          "P0", "✅ Hoàn thành", "Frontend",    "",           "Gallery, biến thể, sticky add-to-cart"),
            ("A4.3", "Quick Add to Cart Modal từ trang danh sách",      "P1", "✅ Hoàn thành", "Frontend",    "",           "Chọn variant và thêm nhanh vào giỏ"),
            ("A4.4", "Giao diện thanh toán 1-page Checkout",            "P0", "✅ Hoàn thành", "Frontend",    "",           "Cart + Checkout tích hợp 1 trang"),
            ("A4.5", "Validate địa chỉ giao hàng – API hành chính VN", "P0", "✅ Hoàn thành", "Frontend",    "",           "Autocomplete Tỉnh/Xã không dấu"),
            ("A4.6", "Áp dụng mã giảm giá (Coupon Code)",              "P0", "✅ Hoàn thành", "Backend",     "",           "Validate hạn, min-order, lượt dùng"),
            ("A4.7", "Thanh toán VietQR – tự sinh mã QR MB Bank",      "P0", "✅ Hoàn thành", "Backend",     "",           "QR chứa số tiền và nội dung CK động"),
            ("A4.8", "Trang theo dõi trạng thái đơn (Order Tracking)", "P0", "✅ Hoàn thành", "Frontend",    "",           "Timeline 4 chiều độc lập"),
            ("A4.9", "Tích hợp cổng thanh toán online (SePay/Stripe)", "P2", "⬜ Chưa làm",  "Backend",     "30/08/2026", "Tự động xác nhận qua API ngân hàng"),
        ]
    },
    {
        "title": "🔐  PHÂN HỆ ĐĂNG KÝ & XÁC THỰC",
        "color": "7C3AED",
        "rows": [
            ("A1.1", "Đăng ký tài khoản bằng Email + xác thực link",   "P0", "✅ Hoàn thành", "Backend",     "",           "pending_registrations + SMTP verification"),
            ("A1.2", "Đăng nhập bằng SĐT + OTP (Stringee/SpeedSMS)",   "P0", "✅ Hoàn thành", "Backend",     "",           "Provider Adapter Pattern, chống spam"),
            ("A1.3", "Đếm ngược OTP 60s – chặn resend & max attempts", "P0", "✅ Hoàn thành", "Backend",     "",           "Rate limit server-side + UI countdown"),
            ("A1.4", "Đăng nhập qua Google / Facebook (OAuth)",         "P1", "✅ Hoàn thành", "Frontend",    "",           "OAuth2 social login integration"),
            ("A1.5", "Duy trì đăng nhập bằng JWT Cookie (HTTP-Only)",   "P0", "✅ Hoàn thành", "Backend",     "",           "Chống XSS/CSRF"),
            ("A1.6", "Tặng 50.000 điểm Welcome Bonus khi đăng ký",     "P0", "✅ Hoàn thành", "Backend",     "",           "Ghi ledger tự động khi kích hoạt tk"),
        ]
    },
    {
        "title": "🤖  PHÂN HỆ TƯ VẤN AI – PET ADVISOR",
        "color": "059669",
        "rows": [
            ("A3.1", "Popup khảo sát thông tin thú cưng (Quiz AI)",     "P0", "✅ Hoàn thành", "Frontend",    "",           "Loài, giống, tuổi, cân nặng, sức khỏe"),
            ("A3.2", "Proxy Groq API an toàn phía Backend",             "P0", "✅ Hoàn thành", "Backend",     "",           "/api/customer/pet-advisor/consult"),
            ("A3.3", "Gợi ý 9 sản phẩm chia 3 nhóm (Budget/Mid/Pro)", "P0", "✅ Hoàn thành", "Backend",     "",           "AI khớp ID sản phẩm thực từ CSDL"),
            ("A3.4", "Kết quả AI: tóm tắt sức khỏe + cảnh báo + voucher","P0","✅ Hoàn thành","Frontend",   "",           "Voucher tích hợp thẻ răng cưa VoucherCard"),
            ("A3.5", "Tự động lưu hồ sơ tư vấn vào CSDL customer_pets","P0", "✅ Hoàn thành", "Backend",     "",           "Gắn với tài khoản khi đã đăng nhập"),
            ("A3.6", "Trang Lịch sử tư vấn dinh dưỡng – Account",     "P0", "✅ Hoàn thành", "Frontend",    "",           "Danh sách + modal chi tiết tư vấn"),
            ("A3.7", "Fallback kết quả khi Groq API gián đoạn",        "P0", "✅ Hoàn thành", "Backend",     "",           "FALLBACK_RESULT có sản phẩm thật"),
        ]
    },
    {
        "title": "⭐  PHÂN HỆ 3F CLUB – LOYALTY & ĐIỂM THƯỞNG",
        "color": "D97706",
        "rows": [
            ("A2.1", "Trang 3F Club – hạng thành viên động (Member→Diamond)","P0","✅ Hoàn thành","Frontend","",  "Rolling 12 tháng, multiplier điểm"),
            ("A2.2", "Thanh tiến trình thăng hạng kép (đơn + chi tiêu)","P0", "✅ Hoàn thành", "Frontend",    "",           "2 KPIs thăng hạng độc lập"),
            ("A2.3", "Lịch sử tích lũy & đổi điểm của khách hàng",     "P0", "✅ Hoàn thành", "Frontend",    "",           "Bảng transactions từ loyalty_point_transactions"),
            ("A2.4", "Trang đổi quà tích điểm (/3f-club/rewards)",     "P0", "✅ Hoàn thành", "Frontend",    "",           "Danh mục quà + yêu cầu đổi + trừ điểm"),
            ("A2.5", "Xác thực SĐT bắt buộc trước khi đổi quà (OTP)", "P1", "✅ Hoàn thành", "Backend",     "",           "OTP gate trước redemption"),
            ("A8.1", "Admin: Duyệt đơn Shopee tích điểm",              "P0", "✅ Hoàn thành", "Backend",     "",           "OCR + Shopee Sandbox API verification"),
            ("A8.2", "Admin: Chia sẻ QR tích điểm cho Sale (Zalo)",    "P1", "✅ Hoàn thành", "Frontend",    "",           "Download PNG QR dẫn đến /3f-club/rewards"),
            ("A8.3", "Admin: Cấu hình hạng thành viên & hệ số điểm",  "P0", "✅ Hoàn thành", "Backend",     "",           "CRUD tiers + lock Diamond"),
            ("A8.4", "Admin: Quản lý Quà đổi điểm & duyệt Redemption","P0", "✅ Hoàn thành", "Backend",     "",           "CRUD rewards, approval workflow"),
            ("A8.5", "Admin: Cộng/trừ điểm thủ công cho thành viên",  "P2", "⬜ Chưa làm",  "Backend",     "15/07/2026", "Phase 6 – ledger manual adjustment"),
            ("A8.6", "OTP xác nhận khi Admin đổi điểm/đổi quà",       "P2", "⬜ Chưa làm",  "Backend",     "31/07/2026", "Phase 6 – bảo mật tài khoản điểm"),
            ("A8.7", "Điểm Holding từ kênh Shopee/TikTok (escrow)",    "P2", "⬜ Chưa làm",  "Backend",     "30/09/2026", "Phase 6 – chờ đơn xác nhận mới unlock"),
            ("A8.8", "Thông báo email nhắc điểm sắp hết hạn",          "P2", "⬜ Chưa làm",  "Backend",     "30/09/2026", "Phase 6 – cron job + SMTP trigger"),
        ]
    },
    {
        "title": "📦  PHÂN HỆ QUẢN LÝ ĐƠN HÀNG (ADMIN)",
        "color": "0891B2",
        "rows": [
            ("A6.1", "Danh sách đơn hàng với bộ lọc đa chỉ số",        "P0", "✅ Hoàn thành", "Frontend",    "",           "Lọc theo: trạng thái, thanh toán, giao nhận, ngày"),
            ("A6.2", "CRM Timeline – nhật ký hoạt động khách hàng",     "P0", "✅ Hoàn thành", "Backend",     "",           "order_status_logs theo 4 trục độc lập"),
            ("A6.3", "Quản lý 4 trạng thái đơn hàng độc lập",          "P0", "✅ Hoàn thành", "Backend",     "",           "Đơn / Thanh toán / Giao hàng / Tích điểm"),
            ("A6.4", "Cấu hình quy trình & bước chuyển trạng thái",    "P0", "✅ Hoàn thành", "Backend",     "",           "Automation rules từ CSDL cấu hình động"),
            ("A6.5", "Thống kê đơn hàng real-time trên Dashboard",     "P0", "✅ Hoàn thành", "Frontend",    "",           "6 KPI cards + biểu đồ Recharts combo"),
        ]
    },
    {
        "title": "📁  PHÂN HỆ QUẢN LÝ CATALOG & SẢN PHẨM (ADMIN)",
        "color": "6366F1",
        "rows": [
            ("A7.1", "CRUD Sản phẩm, Danh mục, Thương hiệu",           "P0", "✅ Hoàn thành", "Backend",     "",           "113 SP + 910 variants đã migration"),
            ("A7.2", "Giao diện Variant Cards trực quan",               "P0", "✅ Hoàn thành", "Frontend",    "",           "Tự sinh SKU, validate trùng, scroll đến lỗi"),
            ("A7.3", "Lưu SP thiếu ảnh dưới dạng Warning (không block)","P1", "✅ Hoàn thành", "Frontend",    "",           "UX không chặn lưu, hiển thị cảnh báo"),
            ("A7.4", "Quản lý Banner Slider trang chủ",                 "P1", "✅ Hoàn thành", "Backend",     "",           "CRUD + lịch chạy + analytics clicks/views"),
            ("A7.5", "Tính năng Wishlist – yêu thích sản phẩm",        "P1", "✅ Hoàn thành", "Backend",     "",           "Guest dùng localStorage, Member dùng CSDL"),
        ]
    },
    {
        "title": "📰  PHÂN HỆ TIN TỨC & BLOG",
        "color": "0891B2",
        "rows": [
            ("A5.1", "Trang danh sách tin tức chuẩn SEO, phân trang",  "P1", "✅ Hoàn thành", "Frontend",    "",           "Lọc danh mục, canonical, sitemap"),
            ("A5.2", "Trang chi tiết bài viết 3 cột – ToC + reading progress","P1","✅ Hoàn thành","Frontend","","JSON-LD structured data, social share bar"),
            ("A5.3", "Cào tin tức tự động từ 3fstore.vn/tin-tuc",      "P1", "✅ Hoàn thành", "Backend",     "",           "Crawler chuyển đổi ảnh tương đối → tuyệt đối"),
            ("A5.4", "Admin: Quản lý bài viết & soạn thảo nội dung",   "P1", "✅ Hoàn thành", "Frontend",    "",           "Rich-text editor, SEO metadata injector"),
        ]
    },
    {
        "title": "🔔  PHÂN HỆ THÔNG BÁO & TRẢI NGHIỆM ADMIN",
        "color": "DC2626",
        "rows": [
            ("A9.1", "Admin Bell – thông báo real-time (polling 60s)", "P1", "✅ Hoàn thành", "Frontend",    "",           "Dropdown 10 thông báo mới nhất + unread count"),
            ("A9.2", "Trigger thông báo: đơn mới, Shopee mới, review mới","P1","✅ Hoàn thành","Backend","",   "Tự động tạo admin_notifications"),
            ("A9.3", "Biểu đồ doanh thu Recharts – so sánh 2 kỳ",     "P1", "✅ Hoàn thành", "Frontend",    "",           "ComposedChart Bar + Line + trục Y đôi"),
            ("A9.4", "Dashboard – Top sản phẩm bán chạy (có filter)",  "P1", "✅ Hoàn thành", "Backend",     "",           "Filter: hôm nay / 7 ngày / 30 ngày"),
            ("A9.5", "Dashboard – Top nhu cầu thú cưng từ AI Advisor", "P1", "✅ Hoàn thành", "Backend",     "",           "Dữ liệu thật từ customer_pets.ai_result"),
            ("A9.6", "Dashboard – Funnel đơn hàng + Top danh mục",     "P2", "⬜ Chưa làm",  "Backend",     "15/07/2026", "4 widget mới: Funnel, Category, Pet Donut, KPI Sparklines"),
        ]
    },
    {
        "title": "🌐  PHÂN HỆ TRANG THÔNG TIN & SEO",
        "color": "047857",
        "rows": [
            ("A10.1","Trang Giới thiệu /about – thiết kế pet store ấm áp","P1","✅ Hoàn thành","Frontend",   "",           "8 sections, SEO JSON-LD, internal links"),
            ("A10.2","Trang Liên hệ /contact – form + Google Maps",     "P1", "✅ Hoàn thành", "Frontend",    "",           "Honeypot chống spam, lưu contact_messages"),
            ("A10.3","Trang Đổi quà 3F Club – giao diện khách hàng",   "P0", "✅ Hoàn thành", "Frontend",    "",           "/3f-club/rewards với banner nhắc OTP"),
            ("A10.4","Trang Tài khoản (Lịch sử đơn, địa chỉ, tư vấn AI)","P0","✅ Hoàn thành","Frontend",   "",           "AccountOverview + tab điều hướng"),
            ("A10.5","SEO Metadata động + JSON-LD cho tất cả trang",    "P1", "✅ Hoàn thành", "Frontend",    "",           "Title, description, og:image per-page"),
        ]
    },
    {
        "title": "🚀  PHÂN HỆ TRIỂN KHAI & DevOps",
        "color": "374151",
        "rows": [
            ("D1.1", "Script deploy backend PHP qua FTP (deploy_ftp.py)","P0","✅ Hoàn thành", "DevOps",      "",           "Bỏ qua file không thay đổi, filter thư mục"),
            ("D1.2", "Backend PHP thuần + MySQL (không framework)",      "P0", "✅ Hoàn thành", "Backend",     "",           "Custom Router, Controller pattern"),
            ("D1.3", "Database migration tự động (ensureSchema)",        "P0", "✅ Hoàn thành", "Backend",     "",           "Tự tạo bảng nếu chưa tồn tại"),
            ("D1.4", "Frontend React + Vite + Tailwind CSS",            "P0", "✅ Hoàn thành", "Frontend",    "",           "Code splitting, lazy loading, WebP"),
            ("D1.5", "Tích hợp thanh toán online SePay/Stripe",         "P2", "⬜ Chưa làm",  "Backend",     "30/08/2026", "Webhook xác nhận tự động, không COD"),
            ("D1.6", "CI/CD Pipeline (GitHub Actions hoặc tương đương)","P2", "⬜ Chưa làm",  "DevOps",      "30/09/2026", "Auto deploy khi push to main"),
        ]
    },
]

# ─── Workbook Setup ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Feature Checklist"

# ─── Row 1: Big Title Banner ─────────────────────────────────────────────────
ws.row_dimensions[1].height = 52
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "3F STORE × 3F CLUB  —  BẢNG KIỂM TRA TÍNH NĂNG DỰ ÁN"
c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=18)
c.fill = fill(C_BRAND_DARK)
c.alignment = align("center", "center", False)
c.border = border(top=no, left=no, right=no, bottom=thick)

# ─── Row 2: Sub-info ─────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 22
ws.merge_cells("A2:I2")
c = ws["A2"]
c.value = f"Cập nhật: {datetime.now().strftime('%d/%m/%Y')}   |   Phiên bản: v2.0   |   Phạm vi: Toàn bộ tính năng chính của dự án"
c.font = Font(name="Calibri", bold=False, color="64748B", size=10, italic=True)
c.fill = fill("F0F6FF")
c.alignment = align("center", "center", False)
c.border = border(top=no, bottom=thin, left=no, right=no)

# ─── Row 3: Legend ───────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 20
ws.merge_cells("A3:I3")
c = ws["A3"]
c.value = "Trạng thái:  ✅ Hoàn thành  |  🔄 Đang làm  |  ⬜ Chưa làm  |  ⏸ Tạm hoãn        Độ ưu tiên:  🔴 P0 = Bắt buộc  |  🟠 P1 = Quan trọng  |  🟣 P2 = Nâng cao"
c.font = Font(name="Calibri", bold=False, color="374151", size=9)
c.fill = fill("FAFBFF")
c.alignment = align("center", "center", False)
c.border = border(top=no, bottom=thick, left=no, right=no)

# ─── Row 4: Column Headers ───────────────────────────────────────────────────
HEADERS = ["MÃ TT", "TÊN TÍNH NĂNG / MÔ TẢ", "ƯU TIÊN", "TRẠNG THÁI", "PHỤ TRÁCH", "DEADLINE", "GHI CHÚ KỸ THUẬT", "% HOÀN THÀNH", "LIÊN KẾT"]
COL_WIDTHS = [10, 52, 11, 17, 13, 14, 48, 14, 16]

ws.row_dimensions[4].height = 32
for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_HEADER_ROW)
    c.alignment = align("center", "center", False)
    c.border = border(top=thick, bottom=thick)
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# ─── Write Data ──────────────────────────────────────────────────────────────
current_row = 5
total_done  = 0
total_all   = 0

for group in GROUPS:
    # Group header row
    ws.row_dimensions[current_row].height = 24
    ws.merge_cells(f"A{current_row}:I{current_row}")
    c = ws[f"A{current_row}"]
    c.value = group["title"]
    c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
    c.fill = fill(group["color"])
    c.alignment = align("left", "center", False)
    c.border = border(top=thick, bottom=thick, left=no, right=no)
    current_row += 1

    for i, row_data in enumerate(group["rows"]):
        code, name, prio, status, owner, deadline, note = row_data
        is_done    = "Hoàn thành" in status
        is_wip     = "Đang làm"   in status
        is_todo    = "Chưa làm"   in status

        total_all += 1
        if is_done:
            total_done += 1
            row_bg      = C_DONE_BG  if i % 2 == 0 else "F0FDF4"
            pct         = "100%"
        elif is_wip:
            row_bg      = C_PENDING_BG if i % 2 == 0 else "FFFBF5"
            pct         = "50%"
        else:
            row_bg      = C_TODO_BG  if i % 2 == 0 else "FFF5F7"
            pct         = "0%"

        ws.row_dimensions[current_row].height = 22

        # Code
        c = ws.cell(row=current_row, column=1, value=code)
        c.font = Font(name="Calibri", bold=True, color="0057E7", size=9)
        c.fill = fill(row_bg); c.alignment = align("center", "center", False)
        c.border = border()

        # Feature name
        c = ws.cell(row=current_row, column=2, value=name)
        c.font = font(bold=is_done, color=C_BRAND_DARK if not is_todo else "6B7280", size=10)
        c.fill = fill(row_bg); c.alignment = align("left", "center", True)
        c.border = border()

        # Priority
        prio_colors = {"P0": (C_PRIO_P0, "FEE2E2"), "P1": (C_PRIO_P1, "FFF3E0"), "P2": (C_PRIO_P2, "EDE9FE")}
        pfg, pbg = prio_colors.get(prio, (C_BRAND_DARK, C_WHITE))
        c = ws.cell(row=current_row, column=3, value=prio)
        c.font = Font(name="Calibri", bold=True, color=pfg, size=9)
        c.fill = fill(pbg); c.alignment = align("center", "center", False)
        c.border = border()

        # Status
        if is_done:
            sfg, sbg = C_DONE_TEXT,    C_DONE_BG
        elif is_wip:
            sfg, sbg = C_PENDING_TEXT, C_PENDING_BG
        else:
            sfg, sbg = C_TODO_TEXT,    C_TODO_BG
        c = ws.cell(row=current_row, column=4, value=status)
        c.font = Font(name="Calibri", bold=True, color=sfg, size=9)
        c.fill = fill(sbg); c.alignment = align("center", "center", False)
        c.border = border()

        # Owner
        c = ws.cell(row=current_row, column=5, value=owner)
        c.font = font(color="374151", size=9)
        c.fill = fill(row_bg); c.alignment = align("center", "center", False)
        c.border = border()

        # Deadline
        c = ws.cell(row=current_row, column=6, value=deadline if deadline else "—")
        c.font = Font(name="Calibri", bold=bool(deadline), color="DC2626" if deadline else "94A3B8", size=9)
        c.fill = fill("FFF5F5" if deadline else row_bg)
        c.alignment = align("center", "center", False)
        c.border = border()

        # Note
        c = ws.cell(row=current_row, column=7, value=note)
        c.font = Font(name="Calibri", color="374151", size=9, italic=True)
        c.fill = fill(row_bg); c.alignment = align("left", "center", True)
        c.border = border()

        # % complete
        c = ws.cell(row=current_row, column=8, value=pct)
        c.font = Font(name="Calibri", bold=True, color=sfg, size=9)
        c.fill = fill(sbg); c.alignment = align("center", "center", False)
        c.border = border()

        # Link placeholder
        c = ws.cell(row=current_row, column=9, value="")
        c.fill = fill(row_bg); c.border = border()

        current_row += 1

    # Spacer after group
    ws.row_dimensions[current_row].height = 6
    ws.merge_cells(f"A{current_row}:I{current_row}")
    ws[f"A{current_row}"].fill = fill("F0F6FF")
    current_row += 1

# ─── Summary Footer ──────────────────────────────────────────────────────────
ws.row_dimensions[current_row].height = 30
ws.merge_cells(f"A{current_row}:D{current_row}")
c = ws[f"A{current_row}"]
pct_overall = round(total_done / total_all * 100) if total_all > 0 else 0
c.value = f"📊  TỔNG TIẾN ĐỘ: {total_done}/{total_all} tính năng hoàn thành  —  {pct_overall}%"
c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
c.fill = fill(C_BRAND_DARK)
c.alignment = align("center", "center", False)
c.border = border(top=thick, bottom=thick, left=no, right=no)

ws.merge_cells(f"E{current_row}:I{current_row}")
c = ws[f"E{current_row}"]
pending_count = total_all - total_done
c.value = f"⬜ Chưa làm: {pending_count}  |  Dự kiến hoàn tất: 30/09/2026"
c.font = Font(name="Calibri", bold=True, color="DC2626", size=10)
c.fill = fill("FFF1F3")
c.alignment = align("center", "center", False)
c.border = border(top=thick, bottom=thick, left=thin, right=no)

# ─── Freeze panes & Filters ──────────────────────────────────────────────────
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:I{current_row - 2}"

# ─── Print settings ──────────────────────────────────────────────────────────
ws.sheet_view.showGridLines = False
ws.page_setup.orientation    = "landscape"
ws.page_setup.fitToPage      = True
ws.page_setup.fitToWidth     = 1
ws.page_setup.paperSize      = 9  # A4
ws.print_title_rows          = "1:4"

# ─── Tab color ───────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = "0057E7"

# ─── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"✅ File Excel đã được tạo thành công!")
print(f"📁 Đường dẫn: {OUTPUT_PATH}")
print(f"📊 Tổng: {total_done}/{total_all} tính năng hoàn thành ({pct_overall}%)")
print(f"⬜ Chưa làm: {total_all - total_done} tính năng")
